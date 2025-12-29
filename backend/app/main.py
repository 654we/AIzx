import csv
import hashlib
import html
import io
import json
import os
import re
import subprocess
import time
import uuid
from datetime import datetime, timedelta

import httpx
from fastapi import Body, Depends, FastAPI, HTTPException, Request, status
from fastapi import UploadFile
from fastapi import Form
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.responses import RedirectResponse
from fastapi.templating import Jinja2Templates
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from starlette.middleware.sessions import SessionMiddleware
from jose import JWTError, jwt
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import Session, sessionmaker

from app import crud, models, schemas
from app.ai_router import build_provider, get_route, provider_ready
from app.news_crawler import crawl_feeds, collect_feed_items
from app.mcp_search import search_news_via_mcp, fetch_mcp_candidates
from app.auth import create_access_token
from app.config import settings
from app.archive_utils import archive_lock, archive_tz, get_last_week_range, parse_published_date
from app.database import init_archive_db, init_db
from app.deps import get_db
from app.mcp.registry import MCPPluginError, test_plugin

from docx import Document
from openpyxl import load_workbook
import xlrd

app = FastAPI(title=settings.app_name)
app.add_middleware(SessionMiddleware, secret_key=settings.admin_session_secret)
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))
scheduler = BackgroundScheduler()
security = HTTPBearer()
archive_engine_cache = None
archive_session_cache = None
archive_url_cache = ""


def _setting_value(db: Session, key: str, default: str) -> str:
    item = crud.get_setting(db, key)
    return item.value if item else default


def _setting_bool(db: Session, key: str, default: str = "false") -> bool:
    return _setting_value(db, key, default).lower() == "true"


def _ensure_setting(db: Session, key: str, default: str) -> None:
    if not crud.get_setting(db, key):
        crud.upsert_setting(db, key, default)


def _setting_optional(db: Session, key: str) -> str:
    item = crud.get_setting(db, key)
    return item.value if item else ""


def get_archive_db_url(db: Session) -> str:
    return _setting_optional(db, "archive_database_url") or settings.archive_database_url


def get_archive_session(db: Session) -> Session | None:
    global archive_engine_cache, archive_session_cache, archive_url_cache
    url = get_archive_db_url(db)
    if not url:
        return None
    if archive_url_cache != url:
        archive_engine_cache = create_engine(url, pool_pre_ping=True)
        archive_session_cache = sessionmaker(autocommit=False, autoflush=False, bind=archive_engine_cache)
        archive_url_cache = url
    return archive_session_cache()


def mask_database_url(url: str) -> str:
    if "://" not in url:
        return url
    prefix, rest = url.split("://", 1)
    if "@" not in rest:
        return url
    creds, host = rest.split("@", 1)
    if ":" in creds:
        user, _ = creds.split(":", 1)
        return f"{prefix}://{user}:******@{host}"
    return f"{prefix}://******@{host}"


def configure_archive_scheduler(db: Session) -> dict:
    _ensure_setting(db, "archive_enabled", "false")
    _ensure_setting(db, "archive_cron", "0 2 * * 1")
    cron_expr = _setting_value(db, "archive_cron", "0 2 * * 1")
    if scheduler.get_job("archive_job"):
        scheduler.remove_job("archive_job")
    scheduler_error = ""
    if _setting_bool(db, "archive_enabled", "false"):
        try:
            scheduler.add_job(
                run_archive_job,
                CronTrigger.from_crontab(cron_expr),
                id="archive_job",
                replace_existing=True,
            )
        except ValueError as exc:
            scheduler_error = f"归档 cron 无效: {exc}"
    if not scheduler.running:
        scheduler.start(paused=False)
    job = scheduler.get_job("archive_job")
    return {"next_run_at": job.next_run_time.isoformat() if job else "", "error": scheduler_error}


def ensure_default_weather_providers(db: Session) -> None:
    if crud.list_weather_providers(db):
        return
    crud.create_weather_provider(
        db,
        name="Open-Meteo",
        provider_type="open-meteo",
        base_url=settings.weather_api_url,
        api_key="",
        timeout_sec=5,
        enabled=True,
        priority=1,
        extra_config=json.dumps({"geo_url": settings.weather_geo_url}, ensure_ascii=False),
    )
    crud.create_weather_provider(
        db,
        name="高德天气",
        provider_type="gaode",
        base_url="https://restapi.amap.com/v3/weather/weatherInfo",
        api_key="",
        timeout_sec=5,
        enabled=False,
        priority=2,
        extra_config=json.dumps({"extensions": "base", "test_location": "上海"}, ensure_ascii=False),
    )


def configure_news_scheduler(db: Session) -> dict:
    _ensure_setting(db, "news_crawler_enabled", "false")
    _ensure_setting(db, "news_crawler_cron", "")
    _ensure_setting(db, "news_crawler_interval_minutes", "30")
    _ensure_setting(db, "news_crawler_limit", "20")
    _ensure_setting(db, "news_target_count", "20")
    _ensure_setting(db, "news_dedupe_max_rounds", "5")
    _ensure_setting(db, "news_dedupe_max_candidates", "200")
    _ensure_setting(db, "news_source_rss", "true")
    _ensure_setting(db, "news_source_mcp", "false")
    _ensure_setting(db, "news_source_feeds", "true")
    cron_expr = _setting_value(db, "news_crawler_cron", "")
    interval_minutes = int(_setting_value(db, "news_crawler_interval_minutes", "30") or 30)
    if scheduler.get_job("news_crawler"):
        scheduler.remove_job("news_crawler")
    scheduler_error = ""
    if cron_expr:
        try:
            scheduler.add_job(
                run_news_crawler,
                CronTrigger.from_crontab(cron_expr),
                id="news_crawler",
                replace_existing=True,
            )
        except ValueError as exc:
            scheduler_error = f"cron 无效，已回退到间隔模式: {exc}"
    if not scheduler.get_job("news_crawler"):
        scheduler.add_job(
            run_news_crawler,
            "interval",
            minutes=interval_minutes,
            id="news_crawler",
            replace_existing=True,
        )
    if not scheduler.running:
        scheduler.start(paused=False)
    job = scheduler.get_job("news_crawler")
    return {
        "next_run_at": job.next_run_time.isoformat() if job else "",
        "error": scheduler_error,
    }


def parse_weekly_cron(cron_expr: str) -> tuple[str, str]:
    parts = cron_expr.split()
    if len(parts) != 5:
        return "", ""
    minute, hour, _, _, weekday = parts
    if weekday == "*":
        return f"{hour.zfill(2)}:{minute.zfill(2)}", "daily"
    return f"{hour.zfill(2)}:{minute.zfill(2)}", weekday


def build_weekly_cron(time_value: str, weekday: str) -> str:
    if not time_value:
        return ""
    hour, minute = time_value.split(":")
    if weekday == "daily":
        return f"{minute} {hour} * * *"
    return f"{minute} {hour} * * {weekday}"


@app.on_event("startup")
def startup_event() -> None:
    init_db()
    db = next(get_db())
    try:
        admin = crud.get_user_by_username(db, "admin")
        if not admin:
            crud.create_user(db, "admin", "admin", is_admin=True)
        ensure_default_weather_providers(db)
        existing_news = db.query(models.NewsItem).count()
        if existing_news == 0:
            crud.create_news_item(
                db,
                title="欢迎使用资讯频道",
                summary="这里展示最新资讯内容，后续将接入订阅与抓取。",
                source="系统",
                url="https://example.com/welcome",
                published_at="2024-01-01T09:00:00+08:00",
                tags=["推荐"],
            )
    finally:
        db.close()
    db = next(get_db())
    try:
        configure_news_scheduler(db)
        init_archive_db()
        configure_archive_scheduler(db)
    finally:
        db.close()


@app.on_event("shutdown")
def shutdown_event() -> None:
    scheduler.shutdown(wait=False)


def fetch_wechat_openid(code: str) -> str:
    if not settings.wechat_appid or not settings.wechat_secret:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="WeChat config missing")
    params = {
        "appid": settings.wechat_appid,
        "secret": settings.wechat_secret,
        "js_code": code,
        "grant_type": "authorization_code",
    }
    try:
        response = httpx.get(settings.wechat_base_url, params=params, timeout=5.0)
        response.raise_for_status()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="WeChat request failed") from exc
    data = response.json()
    if data.get("errcode"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=data.get("errmsg", "WeChat error"))
    openid = data.get("openid")
    if not openid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="WeChat openid missing")
    return openid


def aqi_description(aqi: int) -> str:
    if aqi <= 50:
        return "优"
    if aqi <= 100:
        return "良"
    if aqi <= 150:
        return "轻度"
    if aqi <= 200:
        return "中度"
    if aqi <= 300:
        return "重度"
    return "严重"


def travel_advice_for(condition: str, temp_c: float, aqi: int) -> list[str]:
    advice = []
    if "雨" in condition:
        advice.append("短时有降水，建议随身携带雨具。")
    if temp_c >= 30:
        advice.append("气温偏高，外出注意防晒补水。")
    if temp_c <= 5:
        advice.append("气温较低，外出注意保暖。")
    if aqi > 100:
        advice.append("空气质量一般，敏感人群减少户外活动。")
    if not advice:
        advice.append("天气舒适，适合常规出行安排。")
    return advice[:2]


def travel_advice_from_ai(db: Session, condition: str, temp_c: float, aqi: int) -> list[str]:
    route = get_route(db, "ai_route_weather", "openai")
    if not provider_ready(db, route):
        return travel_advice_for(condition, temp_c, aqi)
    provider = build_provider(route, db)
    prompt = (
        "请根据以下天气信息给出两条简短出行建议，每条不超过20字："
        f"天气={condition}，温度={temp_c}C，AQI={aqi}。"
        "只返回用中文短句，每条建议一行。"
    )
    try:
        text = provider.generate(prompt)
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        return lines[:2] if lines else travel_advice_for(condition, temp_c, aqi)
    except Exception:
        return travel_advice_for(condition, temp_c, aqi)


def _fetch_open_meteo(location: str, geo_url: str, api_url: str, timeout: float) -> schemas.WeatherResponse:
    geo_resp = httpx.get(
        geo_url,
        params={"name": location, "count": 1, "language": "zh", "format": "json"},
        timeout=timeout,
    )
    geo_resp.raise_for_status()
    geo_data = geo_resp.json()
    results = geo_data.get("results") or []
    if not results:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Location not found")
    info = results[0]
    lat = info.get("latitude")
    lon = info.get("longitude")
    name = info.get("name")
    weather_resp = httpx.get(
        api_url,
        params={
            "latitude": lat,
            "longitude": lon,
            "current": "temperature_2m,relative_humidity_2m,wind_speed_10m,weather_code",
            "timezone": "Asia/Shanghai",
        },
        timeout=timeout,
    )
    weather_resp.raise_for_status()
    weather_data = weather_resp.json()
    current = weather_data.get("current") or {}
    temp_c = float(current.get("temperature_2m", 0))
    humidity = int(current.get("relative_humidity_2m", 0))
    wind_speed = current.get("wind_speed_10m", 0)
    condition = "晴"
    weather_code = current.get("weather_code")
    if weather_code in {51, 53, 55, 61, 63, 65, 80, 81, 82}:
        condition = "雨"
    elif weather_code in {71, 73, 75, 85, 86}:
        condition = "雪"
    elif weather_code in {2, 3}:
        condition = "多云"
    aqi = 60
    aqi_desc = aqi_description(aqi)
    return schemas.WeatherResponse(
        location=schemas.WeatherLocation(name=name, lat=lat, lon=lon),
        weather=schemas.WeatherInfo(
            condition=condition,
            temp_c=temp_c,
            humidity=humidity,
            wind=f"{wind_speed} km/h",
            aqi=aqi,
            aqi_desc=aqi_desc,
            updated_at=current.get("time", ""),
        ),
        travel_advice=[],
    )


def _fetch_gaode(location: str, base_url: str, api_key: str, timeout: float) -> schemas.WeatherResponse:
    if not api_key:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Gaode api_key missing")
    response = httpx.get(
        base_url,
        params={"key": api_key, "city": location, "extensions": "base"},
        timeout=timeout,
    )
    response.raise_for_status()
    data = response.json()
    lives = data.get("lives") or []
    if not lives:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Location not found")
    info = lives[0]
    condition = info.get("weather", "晴")
    temp_c = float(info.get("temperature", 0))
    humidity = int(info.get("humidity", 0))
    wind = f"{info.get('winddirection', '')}{info.get('windpower', '')}"
    aqi = 60
    aqi_desc = aqi_description(aqi)
    return schemas.WeatherResponse(
        location=schemas.WeatherLocation(name=info.get("city", location), lat=0.0, lon=0.0),
        weather=schemas.WeatherInfo(
            condition=condition,
            temp_c=temp_c,
            humidity=humidity,
            wind=wind,
            aqi=aqi,
            aqi_desc=aqi_desc,
            updated_at=info.get("reporttime", ""),
        ),
        travel_advice=[],
    )


def fetch_weather(db: Session, location: str) -> schemas.WeatherResponse:
    if not location:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Location required")
    providers = [
        provider for provider in crud.list_weather_providers(db) if provider.enabled
    ]
    errors = []
    for provider in providers:
        try:
            extra = json.loads(provider.extra_config or "{}")
            if provider.provider_type == "open-meteo":
                response = _fetch_open_meteo(
                    location,
                    geo_url=extra.get("geo_url", settings.weather_geo_url),
                    api_url=provider.base_url,
                    timeout=float(provider.timeout_sec or 5),
                )
            elif provider.provider_type == "gaode":
                response = _fetch_gaode(
                    location,
                    base_url=provider.base_url,
                    api_key=provider.api_key,
                    timeout=float(provider.timeout_sec or 5),
                )
            else:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported provider")
            response.travel_advice = travel_advice_from_ai(
                db,
                response.weather.condition,
                response.weather.temp_c,
                response.weather.aqi,
            )
            return response
        except Exception as exc:
            errors.append(f"{provider.provider_type}:{exc}")
            continue
    if not providers:
        try:
            response = _fetch_open_meteo(
                location,
                geo_url=settings.weather_geo_url,
                api_url=settings.weather_api_url,
                timeout=5.0,
            )
            response.travel_advice = travel_advice_from_ai(
                db,
                response.weather.condition,
                response.weather.temp_c,
                response.weather.aqi,
            )
            return response
        except Exception as exc:
            errors.append(f"open-meteo:{exc}")
    raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="Weather providers failed: " + " | ".join(errors))


def build_schedule_plan(source_filename: str) -> schemas.ScheduleResponse:
    today = datetime.now()
    week = []
    for offset in range(5):
        day = today + timedelta(days=offset)
        date_str = day.strftime("%Y-%m-%d")
        week.append(
            schemas.ScheduleDay(
                date=date_str,
                day_of_week=day.isoweekday(),
                blocks=[
                    schemas.ScheduleBlock(
                        start="09:00",
                        end="10:30",
                        title="待办事项",
                        location="待定",
                        type="other",
                        notes="示例日程，上传内容解析将在后续阶段完善。",
                    )
                ],
            )
        )
    return schemas.ScheduleResponse(
        meta=schemas.ScheduleMeta(
            source_filename=source_filename,
            generated_at=datetime.now().isoformat(),
            timezone="Asia/Shanghai",
        ),
        week=week,
        tips=["上传更多日程内容以生成更准确的计划。"],
    )


def normalize_schedule(plan: schemas.ScheduleResponse) -> schemas.ScheduleResponse:
    normalized_week = []
    for day in plan.week:
        blocks = sorted(day.blocks, key=lambda item: item.start)
        cleaned_blocks = []
        last_end = ""
        for block in blocks:
            if block.end <= block.start:
                continue
            if last_end and block.start < last_end:
                continue
            cleaned_blocks.append(block)
            last_end = block.end
        normalized_week.append(
            schemas.ScheduleDay(
                date=day.date,
                day_of_week=day.day_of_week,
                blocks=cleaned_blocks,
            )
        )
    return schemas.ScheduleResponse(meta=plan.meta, week=normalized_week, tips=plan.tips)


def parse_text_file(content: bytes) -> str:
    return content.decode("utf-8", errors="ignore")


def parse_docx(content: bytes) -> str:
    document = Document(io.BytesIO(content))
    return "\n".join([para.text for para in document.paragraphs if para.text.strip()])


def parse_xlsx(content: bytes) -> str:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) for cell in row if cell is not None]
            if row_values:
                lines.append(" ".join(row_values))
    return "\n".join(lines)


def parse_xls(content: bytes) -> str:
    workbook = xlrd.open_workbook(file_contents=content)
    lines = []
    for sheet in workbook.sheets():
        for row_index in range(sheet.nrows):
            row_values = [str(cell.value) for cell in sheet.row(row_index) if cell.value]
            if row_values:
                lines.append(" ".join(row_values))
    return "\n".join(lines)


def parse_csv(content: bytes) -> str:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        row_values = [item for item in row if item]
        if row_values:
            lines.append(" ".join(row_values))
    return "\n".join(lines)


def schedule_from_ai(db: Session, source_filename: str, content: str) -> schemas.ScheduleResponse:
    route = get_route(db, "ai_route_schedule", "glm")
    if not provider_ready(db, route):
        return build_schedule_plan(source_filename)
    provider = build_provider(route, db)
    prompt = (
        "请根据以下日程内容生成固定格式JSON，不要包含多余文字。"
        "JSON格式：{\"meta\": {\"source_filename\": \"\", \"generated_at\": \"ISO8601\", \"timezone\": \"Asia/Shanghai\", \"version\": \"1.0\"},"
        "\"week\": [{\"date\": \"YYYY-MM-DD\", \"day_of_week\": 1, \"blocks\": [{\"start\": \"HH:MM\", \"end\": \"HH:MM\","
        "\"title\": \"\", \"location\": \"\", \"type\": \"study|work|life|health|other\", \"notes\": \"\"}]}],"
        "\"tips\": [\"\", \"\"]}"
        "确保 blocks 时间不重叠，按 start 升序。内容如下：\n"
        f"{content}"
    )
    try:
        raw = provider.generate(prompt)
        data = json.loads(raw)
        return schemas.ScheduleResponse(**data)
    except Exception:
        return build_schedule_plan(source_filename)


def news_summary_from_ai(db: Session, title: str, content: str) -> str:
    route = get_route(db, "ai_route_news", "deepseek")
    if not provider_ready(db, route):
        return content[:120] if content else "暂无摘要"
    provider = build_provider(route, db)
    prompt = (
        "请为以下资讯生成不超过60字的中文摘要："
        f"标题：{title}\n内容：{content}\n"
        "仅返回摘要文本。"
    )
    try:
        text = provider.generate(prompt)
        return text.strip()[:120] if text else "暂无摘要"
    except Exception:
        return content[:120] if content else "暂无摘要"


def news_keypoints_from_ai(db: Session, title: str, content: str) -> list[str]:
    route = get_route(db, "ai_route_news", "deepseek")
    if not provider_ready(db, route):
        fallback = [line.strip() for line in content.split("。") if line.strip()]
        return fallback[:3] if fallback else ["暂无要点"]
    provider = build_provider(route, db)
    prompt = (
        "请为以下资讯生成3条关键要点，每条不超过20字，使用中文。"
        f"标题：{title}\n内容：{content}\n"
        "仅返回要点列表，每行一条。"
    )
    try:
        text = provider.generate(prompt)
        lines = [line.strip("- ").strip() for line in text.splitlines() if line.strip()]
        return lines[:3] if lines else ["暂无要点"]
    except Exception:
        fallback = [line.strip() for line in content.split("。") if line.strip()]
        return fallback[:3] if fallback else ["暂无要点"]


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.lower().strip())


def compute_content_hash(title: str, summary: str, url: str) -> str:
    base = normalize_text(title) + "|" + normalize_text(summary) + "|" + normalize_text(url)
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def choose_best_candidate(db: Session, candidates: list[dict]) -> tuple[dict, str]:
    if len(candidates) == 1:
        return candidates[0], "唯一候选"
    route = get_route(db, "ai_route_news", "deepseek")
    if not provider_ready(db, route):
        candidates_sorted = sorted(
            candidates,
            key=lambda item: (len(item.get("summary", "")), item.get("published_at", "")),
            reverse=True,
        )
        return candidates_sorted[0], "摘要更完整且发布时间更晚"
    provider = build_provider(route, db)
    prompt_items = []
    for idx, item in enumerate(candidates, start=1):
        prompt_items.append(
            f"{idx}. 标题：{item.get('title')}\n"
            f"摘要：{item.get('summary')}\n"
            f"来源：{item.get('source')}\n"
            f"发布时间：{item.get('published_at')}\n"
            f"链接：{item.get('url')}"
        )
    prompt = (
        "你是资讯去重评审，请基于信息完整度、时效性、来源可信度、标题与正文一致性、结构清晰度，"
        "从候选中选出最佳的一篇。只返回序号与一句理由，格式：\"序号|理由\"。\n"
        + "\n\n".join(prompt_items)
    )
    try:
        result = provider.generate(prompt).strip()
        parts = result.split("|", 1)
        index = int(parts[0].strip()) - 1
        if 0 <= index < len(candidates):
            reason = parts[1].strip() if len(parts) > 1 else "AI评估"
            return candidates[index], reason
    except Exception:
        pass
    candidates_sorted = sorted(
        candidates,
        key=lambda item: (len(item.get("summary", "")), item.get("published_at", "")),
        reverse=True,
    )
    return candidates_sorted[0], "摘要更完整且发布时间更晚"


def extract_text_from_html(raw_html: str) -> str:
    cleaned = re.sub(r"<script.*?>.*?</script>", " ", raw_html, flags=re.S | re.I)
    cleaned = re.sub(r"<style.*?>.*?</style>", " ", cleaned, flags=re.S | re.I)
    cleaned = re.sub(r"<[^>]+>", " ", cleaned)
    cleaned = html.unescape(cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def extract_title_from_html(raw_html: str, fallback: str) -> str:
    match = re.search(r"<title>(.*?)</title>", raw_html, flags=re.I | re.S)
    if match:
        return html.unescape(match.group(1)).strip()
    return fallback


def build_news_preview(
    db: Session,
    source_url: str,
    title_hint: str,
    content_text: str,
    cache_ttl_sec: int,
    news_id: int | None,
) -> schemas.NewsPreviewResponse:
    summary = news_summary_from_ai(db, title_hint, content_text)
    key_points = news_keypoints_from_ai(db, title_hint, content_text)
    fetched_at = datetime.now().isoformat()
    crud.upsert_news_preview(
        db,
        news_id=news_id,
        source_url=source_url,
        title=title_hint,
        summary=summary,
        key_points=json.dumps(key_points, ensure_ascii=False),
        fetched_at=fetched_at,
        cache_ttl_sec=cache_ttl_sec,
    )
    return schemas.NewsPreviewResponse(
        title=title_hint,
        summary=summary,
        key_points=key_points,
        source_url=source_url,
        fetched_at=fetched_at,
    )


def run_news_crawler(limit: int | None = None, force_run: bool = False) -> dict:
    db = next(get_db())
    created = 0
    used_sources = []
    try:
        enabled = _setting_bool(db, "news_crawler_enabled", "false")
        if not enabled and not force_run:
            return {"created": 0, "sources": []}
        sources = {
            "rss": _setting_bool(db, "news_source_rss", "true"),
            "mcp": _setting_bool(db, "news_source_mcp", "false"),
            "feeds": _setting_bool(db, "news_source_feeds", "true"),
        }
        target_count = int(_setting_value(db, "news_target_count", "20") or 20)
        if limit is not None:
            target_count = min(target_count, limit)
        max_rounds = int(_setting_value(db, "news_dedupe_max_rounds", "5") or 5)
        max_candidates = int(_setting_value(db, "news_dedupe_max_candidates", "200") or 200)
        feed_setting = crud.get_setting(db, "news_feed_urls")
        feeds = [item.strip() for item in (feed_setting.value.split(",") if feed_setting and feed_setting.value else []) if item.strip()]
        candidates: list[dict] = []
        round_count = 0
        while len(candidates) < max_candidates and round_count < max_rounds:
            round_count += 1
            round_limit = max_candidates - len(candidates)
            if sources["mcp"]:
                candidates.extend(fetch_mcp_candidates(db, min(10, round_limit)))
                used_sources.append("mcp")
            if feeds and (sources["rss"] or sources["feeds"]):
                candidates.extend(collect_feed_items(feeds, limit=min(20, round_limit)))
                used_sources.append("rss")
            if len(candidates) >= max_candidates:
                break
        groups: list[list[dict]] = []
        for item in candidates:
            text = normalize_text(item.get("title", "") + " " + item.get("summary", ""))
            matched = False
            for group in groups:
                base_text = normalize_text(group[0].get("title", "") + " " + group[0].get("summary", ""))
                if base_text and (base_text == text or (len(base_text) > 20 and base_text in text) or (len(text) > 20 and text in base_text)):
                    group.append(item)
                    matched = True
                    break
            if not matched:
                groups.append([item])
        deduped = []
        for group in groups:
            best, reason = choose_best_candidate(db, group)
            merged_urls = [item["url"] for item in group if item["url"] != best["url"]]
            best["dedupe_keep_reason"] = reason
            best["dedupe_merged_urls"] = json.dumps(merged_urls, ensure_ascii=False)
            best["dedupe_group_id"] = str(uuid.uuid4())
            deduped.append(best)
        deduped = deduped[:target_count]
        for item in deduped:
            if crud.get_news_by_url(db, item["url"]):
                continue
            summary = item.get("summary") or news_summary_from_ai(db, item.get("title", ""), "")
            content_hash = compute_content_hash(item.get("title", ""), summary, item.get("url", ""))
            crud.create_news_item(
                db,
                title=item.get("title", "未命名资讯"),
                summary=summary[:200],
                source=item.get("source", "订阅源"),
                url=item.get("url", ""),
                published_at=item.get("published_at", datetime.now().isoformat()),
                tags=item.get("tags", ["订阅"]),
                content_hash=content_hash,
                dedupe_group_id=item.get("dedupe_group_id", ""),
                dedupe_keep_reason=item.get("dedupe_keep_reason", ""),
                dedupe_merged_urls=item.get("dedupe_merged_urls", "[]"),
            )
            created += 1
    finally:
        db.close()
    return {"created": created, "sources": used_sources}


def run_archive_job(force_run: bool = False) -> dict:
    if not archive_lock.acquire(blocking=False):
        return {"archived": 0, "deleted": 0, "status": "running"}
    db = next(get_db())
    archive_session = None
    start_time = time.time()
    status_text = "success"
    error_message = ""
    archived_count = 0
    deleted_count = 0
    week_key = ""
    try:
        enabled = _setting_bool(db, "archive_enabled", "false")
        if not enabled and not force_run:
            return {"archived": 0, "deleted": 0, "status": "disabled"}
        archive_session = get_archive_session(db)
        if not archive_session:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="ARCHIVE_DB_NOT_CONFIGURED")
        start, end, week_key = get_last_week_range()
        items = db.query(models.NewsItem).all()
        to_archive = []
        for item in items:
            published_date = parse_published_date(item.published_at)
            if not published_date:
                continue
            if start <= published_date <= end:
                to_archive.append(item)
        for item in to_archive:
            exists = (
                archive_session.query(models.ArchiveNewsItem)
                .filter(models.ArchiveNewsItem.url == item.url, models.ArchiveNewsItem.archive_week == week_key)
                .first()
            )
            if exists:
                continue
            archive_session.add(
                models.ArchiveNewsItem(
                    title=item.title,
                    summary=item.summary,
                    source=item.source,
                    url=item.url,
                    published_at=item.published_at,
                    tags=item.tags,
                    content_hash=item.content_hash,
                    archive_week=week_key,
                )
            )
            archived_count += 1
        archive_session.commit()
        for item in to_archive:
            db.delete(item)
            deleted_count += 1
        db.commit()
    except Exception as exc:
        if archive_session:
            archive_session.rollback()
        db.rollback()
        status_text = "failed"
        error_message = str(exc)
    finally:
        if archive_session:
            archive_session.close()
        db.close()
        archive_lock.release()
    duration_ms = int((time.time() - start_time) * 1000)
    task_db = next(get_db())
    try:
        crud.create_task_run(
            task_db,
            task_type="archive_weekly",
            status=status_text,
            duration_ms=duration_ms,
            error_message=error_message,
            log_excerpt=json.dumps({"week": week_key, "archived": archived_count, "deleted": deleted_count}, ensure_ascii=False),
            payload_json=json.dumps({"week": week_key}, ensure_ascii=False),
        )
    finally:
        task_db.close()
    return {"archived": archived_count, "deleted": deleted_count, "week": week_key, "status": status_text}


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, settings.jwt_secret, algorithms=[settings.jwt_algorithm])
        username: str | None = payload.get("sub")
        if not username:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    except JWTError as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token") from exc
    user = crud.get_user_by_username(db, username)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="User disabled")
    return user


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok"}


@app.post("/api/auth/register", response_model=schemas.Token)
def register(payload: schemas.UserCreate, db: Session = Depends(get_db)):
    existing = crud.get_user_by_username(db, payload.username)
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Username already exists")
    user = crud.create_user(db, payload.username, payload.password)
    token = create_access_token(user.username)
    return {"access_token": token, "token_type": "bearer"}


@app.post("/api/auth/login", response_model=schemas.Token)
def login(payload: schemas.UserLogin, db: Session = Depends(get_db)):
    user = crud.authenticate_user(db, payload.username, payload.password)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
    token = create_access_token(user.username)
    return {"access_token": token, "token_type": "bearer"}


@app.post("/api/auth/wechat_login", response_model=schemas.Token)
def wechat_login(payload: schemas.WechatLogin, db: Session = Depends(get_db)):
    openid = fetch_wechat_openid(payload.code)
    user = crud.get_user_by_openid(db, openid)
    if not user:
        username = f"wx_{openid[:16]}"
        user = crud.create_wechat_user(db, username, openid)
    token = create_access_token(user.username)
    return {"access_token": token, "token_type": "bearer"}


@app.get("/api/weather", response_model=schemas.WeatherResponse)
def weather(location: str, db: Session = Depends(get_db)):
    return fetch_weather(db, location)


@app.get("/api/news", response_model=schemas.NewsResponse)
def news(
    page: int = 1,
    page_size: int = 10,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    page = max(page, 1)
    page_size = min(max(page_size, 1), 50)
    tags = [tag for tag in current_user.subscriptions.split(",") if tag]
    items, total = crud.list_news(db, tags, page, page_size)
    mapped = [
        schemas.NewsItem(
            id=item.id,
            title=item.title,
            summary=item.summary,
            source=item.source,
            url=item.url,
            published_at=item.published_at,
            tags=[tag for tag in item.tags.split(",") if tag],
        )
        for item in items
    ]
    return schemas.NewsResponse(
        items=mapped,
        page=page,
        page_size=page_size,
        has_more=page * page_size < total,
    )


@app.get("/api/news/{news_id}/preview", response_model=schemas.NewsPreviewResponse)
def news_preview(news_id: int, db: Session = Depends(get_db)):
    preview = crud.get_news_preview_by_news_id(db, news_id)
    if preview:
        try:
            fetched_time = datetime.fromisoformat(preview.fetched_at)
            if (datetime.now() - fetched_time).total_seconds() < preview.cache_ttl_sec:
                return schemas.NewsPreviewResponse(
                    title=preview.title,
                    summary=preview.summary,
                    key_points=json.loads(preview.key_points or "[]"),
                    source_url=preview.source_url,
                    fetched_at=preview.fetched_at,
                )
        except Exception:
            pass
    item = db.query(models.NewsItem).filter(models.NewsItem.id == news_id).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="News not found")
    try:
        response = httpx.get(item.url, timeout=8.0)
        response.raise_for_status()
        title_hint = extract_title_from_html(response.text, item.title)
        text = extract_text_from_html(response.text)[:2000]
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="Preview fetch failed") from exc
    return build_news_preview(
        db,
        source_url=item.url,
        title_hint=title_hint,
        content_text=text,
        cache_ttl_sec=3600,
        news_id=news_id,
    )


@app.post("/api/news/preview", response_model=schemas.NewsPreviewResponse)
def news_preview_by_url(payload: dict = Body(default={}), db: Session = Depends(get_db)):
    url = payload.get("url", "")
    if not url:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="url required")
    preview = crud.get_news_preview_by_url(db, url)
    if preview:
        try:
            fetched_time = datetime.fromisoformat(preview.fetched_at)
            if (datetime.now() - fetched_time).total_seconds() < preview.cache_ttl_sec:
                return schemas.NewsPreviewResponse(
                    title=preview.title,
                    summary=preview.summary,
                    key_points=json.loads(preview.key_points or "[]"),
                    source_url=preview.source_url,
                    fetched_at=preview.fetched_at,
                )
        except Exception:
            pass
    try:
        response = httpx.get(url, timeout=8.0)
        response.raise_for_status()
        title_hint = extract_title_from_html(response.text, url)
        text = extract_text_from_html(response.text)[:2000]
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail="Preview fetch failed") from exc
    return build_news_preview(
        db,
        source_url=url,
        title_hint=title_hint,
        content_text=text,
        cache_ttl_sec=3600,
        news_id=None,
    )


@app.get("/api/archive/weeks")
def archive_weeks(
    last_n_weeks: int = 8,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    archive_session = get_archive_session(db)
    if not archive_session:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="ARCHIVE_DB_NOT_CONFIGURED")
    try:
        last_n_weeks = min(max(last_n_weeks, 1), 52)
        rows = (
            archive_session.query(
                models.ArchiveNewsItem.archive_week,
                func.count(models.ArchiveNewsItem.id),
            )
            .group_by(models.ArchiveNewsItem.archive_week)
            .order_by(models.ArchiveNewsItem.archive_week.desc())
            .limit(last_n_weeks)
            .all()
        )
        week_counts = {week: count for week, count in rows}
        sorted_weeks = [week for week, _ in rows]
    finally:
        archive_session.close()
    return {
        "weeks": sorted_weeks,
        "default": sorted_weeks[0] if sorted_weeks else "",
        "counts": {week: week_counts[week] for week in sorted_weeks},
    }


@app.get("/api/archive/news", response_model=schemas.NewsResponse)
def archive_news(
    week: str | None = None,
    week_start: str | None = None,
    week_end: str | None = None,
    page: int = 1,
    page_size: int = 10,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    archive_session = get_archive_session(db)
    if not archive_session:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="ARCHIVE_DB_NOT_CONFIGURED")
    page = max(page, 1)
    page_size = min(max(page_size, 1), 50)
    query = archive_session.query(models.ArchiveNewsItem)
    if week:
        query = query.filter(models.ArchiveNewsItem.archive_week == week)
    if week_start:
        query = query.filter(models.ArchiveNewsItem.published_at >= week_start)
    if week_end:
        query = query.filter(models.ArchiveNewsItem.published_at <= week_end)
    total = query.count()
    items = (
        query.order_by(models.ArchiveNewsItem.published_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    mapped = [
        schemas.NewsItem(
            id=item.id,
            title=item.title,
            summary=item.summary,
            source=item.source,
            url=item.url,
            published_at=item.published_at,
            tags=[tag for tag in item.tags.split(",") if tag],
        )
        for item in items
    ]
    archive_session.close()
    stats = {"count": total}
    return schemas.NewsResponse(
        items=mapped,
        page=page,
        page_size=page_size,
        has_more=page * page_size < total,
        stats=stats,
    )


@app.post("/api/schedule/upload", response_model=schemas.ScheduleResponse)
def upload_schedule(
    file: UploadFile,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File required")
    if not file.filename.lower().endswith((".txt", ".md", ".csv", ".doc", ".docx", ".xls", ".xlsx")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file type")
    content = file.file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File too large")
    os.makedirs("uploads", exist_ok=True)
    stored_path = os.path.join("uploads", f"{current_user.id}_{file.filename}")
    with open(stored_path, "wb") as f:
        f.write(content)
    file_lower = file.filename.lower()
    if file_lower.endswith((".txt", ".md")):
        content_text = parse_text_file(content)
    elif file_lower.endswith(".csv"):
        content_text = parse_csv(content)
    elif file_lower.endswith(".docx"):
        content_text = parse_docx(content)
    elif file_lower.endswith(".doc"):
        content_text = parse_text_file(content)
    elif file_lower.endswith(".xlsx"):
        content_text = parse_xlsx(content)
    elif file_lower.endswith(".xls"):
        content_text = parse_xls(content)
    else:
        content_text = ""
    crud.create_schedule_upload(
        db,
        user_id=current_user.id,
        filename=file.filename,
        file_type=os.path.splitext(file.filename)[-1].lstrip("."),
        file_size=len(content),
        stored_path=stored_path,
        status="parsed" if content_text else "failed",
        parsed_text=content_text,
        created_at=datetime.now().isoformat(),
    )
    plan = schedule_from_ai(db, file.filename, content_text)
    plan = normalize_schedule(plan)
    payload = plan.model_dump_json()
    crud.create_schedule_plan(
        db,
        user_id=current_user.id,
        source_filename=file.filename,
        payload=payload,
        created_at=datetime.now().isoformat(),
    )
    return plan


@app.get("/api/schedule/latest", response_model=schemas.ScheduleResponse)
def latest_schedule(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    plan = crud.get_latest_schedule(db, current_user.id)
    if not plan:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No schedule found")
    return schemas.ScheduleResponse(**json.loads(plan.payload))


@app.get("/api/user/profile", response_model=schemas.UserProfile)
def profile(current_user=Depends(get_current_user)):
    subscriptions = [tag for tag in current_user.subscriptions.split(",") if tag]
    return schemas.UserProfile(
        id=current_user.id,
        username=current_user.username,
        location=current_user.location,
        email=current_user.email,
        phone=current_user.phone,
        avatar_url=current_user.avatar_url,
        subscriptions=subscriptions,
    )


@app.put("/api/user/profile")
def update_profile(
    payload: schemas.UpdateUserProfile,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    existing = crud.get_user_by_username(db, payload.username)
    if existing and existing.id != current_user.id:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Username already exists")
    crud.update_user_profile(
        db,
        current_user,
        username=payload.username,
        location=payload.location,
        subscriptions=[tag for tag in current_user.subscriptions.split(",") if tag],
        email=payload.email,
        phone=payload.phone,
        avatar_url=payload.avatar_url,
    )
    return {"status": "ok"}


@app.put("/api/user/location")
def update_location(
    payload: schemas.UpdateLocation,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    current_user.location = payload.location
    db.add(current_user)
    db.commit()
    return {"status": "ok"}


@app.put("/api/user/subscriptions")
def update_subscriptions(
    payload: schemas.UpdateSubscriptions,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    current_user.subscriptions = ",".join(payload.tags)
    db.add(current_user)
    db.commit()
    return {"status": "ok"}


@app.get("/api/user/subscriptions")
def get_subscriptions(current_user=Depends(get_current_user)):
    return {"tags": [tag for tag in current_user.subscriptions.split(",") if tag]}


@app.put("/api/user/password")
def update_password(
    payload: schemas.UpdateUserPassword,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    user = crud.authenticate_user(db, current_user.username, payload.old_password)
    if not user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Old password incorrect")
    crud.reset_user_password(db, current_user, payload.new_password)
    return {"status": "ok"}


@app.get("/admin/login")
def admin_login(request: Request):
    return templates.TemplateResponse("admin_login.html", {"request": request, "error": ""})


@app.post("/admin/login")
def admin_login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = crud.authenticate_user(db, username, password)
    if not user or not user.is_admin:
        return templates.TemplateResponse(
            "admin_login.html",
            {"request": request, "error": "账号或密码错误"},
        )
    request.session["admin_user"] = user.username
    return RedirectResponse(url="/admin/settings", status_code=status.HTTP_302_FOUND)


def require_admin(request: Request):
    if not request.session.get("admin_user"):
        return RedirectResponse(url="/admin/login", status_code=status.HTTP_302_FOUND)
    return None


def require_admin_json(request: Request) -> None:
    if not request.session.get("admin_user"):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Admin required")


@app.get("/admin/logout")
def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/admin/login", status_code=status.HTTP_302_FOUND)


def _task_run_to_dict(run: models.TaskRun) -> dict:
    return {
        "id": run.id,
        "task_type": run.task_type,
        "status": run.status,
        "duration_ms": run.duration_ms,
        "error_message": run.error_message,
        "log_excerpt": run.log_excerpt,
        "payload_json": run.payload_json,
        "created_at": run.created_at,
    }


@app.get("/admin/settings")
def admin_settings(request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    def setting_value(key: str, default: str) -> str:
        item = crud.get_setting(db, key)
        return item.value if item else default

    settings_map = {
        "wechat_appid": setting_value("wechat_appid", ""),
        "wechat_secret": setting_value("wechat_secret", ""),
        "weather_geo_url": setting_value("weather_geo_url", settings.weather_geo_url),
        "weather_api_url": setting_value("weather_api_url", settings.weather_api_url),
        "news_crawler_enabled": setting_value("news_crawler_enabled", "false"),
        "news_crawler_cron": setting_value("news_crawler_cron", ""),
        "news_crawler_interval_minutes": setting_value("news_crawler_interval_minutes", "30"),
        "news_crawler_limit": setting_value("news_crawler_limit", "20"),
        "news_target_count": setting_value("news_target_count", "20"),
        "news_dedupe_max_rounds": setting_value("news_dedupe_max_rounds", "5"),
        "news_dedupe_max_candidates": setting_value("news_dedupe_max_candidates", "200"),
        "news_source_rss": setting_value("news_source_rss", "true"),
        "news_source_mcp": setting_value("news_source_mcp", "false"),
        "news_source_feeds": setting_value("news_source_feeds", "true"),
        "news_feed_urls": setting_value("news_feed_urls", ""),
        "news_source_mode": setting_value("news_source_mode", "rss"),
        "mcp_enabled": setting_value("mcp_enabled", "false"),
        "mcp_base_url": setting_value("mcp_base_url", ""),
        "mcp_api_key": setting_value("mcp_api_key", ""),
        "mcp_keywords": setting_value("mcp_keywords", "资讯,科技,天气"),
        "ai_deepseek_api_key": setting_value("ai_deepseek_api_key", ""),
        "ai_deepseek_base_url": setting_value("ai_deepseek_base_url", "https://api.deepseek.com"),
        "ai_deepseek_model": setting_value("ai_deepseek_model", "deepseek-chat"),
        "ai_glm_api_key": setting_value("ai_glm_api_key", ""),
        "ai_glm_base_url": setting_value("ai_glm_base_url", "https://open.bigmodel.cn/api/paas/v4"),
        "ai_glm_model": setting_value("ai_glm_model", "glm-4"),
        "ai_openai_api_key": setting_value("ai_openai_api_key", ""),
        "ai_openai_base_url": setting_value("ai_openai_base_url", "https://api.openai.com/v1"),
        "ai_openai_model": setting_value("ai_openai_model", "gpt-4o-mini"),
        "ai_route_news": setting_value("ai_route_news", "deepseek"),
        "ai_route_weather": setting_value("ai_route_weather", "openai"),
        "ai_route_schedule": setting_value("ai_route_schedule", "glm"),
        "archive_database_url": mask_database_url(setting_value("archive_database_url", settings.archive_database_url)),
        "archive_enabled": setting_value("archive_enabled", "false"),
        "archive_cron": setting_value("archive_cron", "0 2 * * 1"),
    }
    news_time, news_weekday = parse_weekly_cron(settings_map["news_crawler_cron"])
    archive_time, archive_weekday = parse_weekly_cron(settings_map["archive_cron"])
    settings_map["news_run_time"] = news_time
    settings_map["news_run_weekday"] = news_weekday
    settings_map["archive_run_time"] = archive_time
    settings_map["archive_run_weekday"] = archive_weekday
    return templates.TemplateResponse(
        "admin_settings.html",
        {"request": request, "settings": settings_map},
    )


@app.post("/admin/settings")
def admin_settings_post(
    request: Request,
    wechat_appid: str = Form(""),
    wechat_secret: str = Form(""),
    weather_geo_url: str = Form(""),
    weather_api_url: str = Form(""),
    news_crawler_enabled: str = Form("false"),
    news_crawler_cron: str = Form(""),
    news_crawler_interval_minutes: str = Form("30"),
    news_crawler_limit: str = Form("20"),
    news_target_count: str = Form("20"),
    news_dedupe_max_rounds: str = Form("5"),
    news_dedupe_max_candidates: str = Form("200"),
    news_run_time: str = Form(""),
    news_run_weekday: str = Form(""),
    news_source_rss: str = Form("true"),
    news_source_mcp: str = Form("false"),
    news_source_feeds: str = Form("true"),
    news_feed_urls: str = Form(""),
    news_source_mode: str = Form("rss"),
    mcp_enabled: str = Form("false"),
    mcp_base_url: str = Form(""),
    mcp_api_key: str = Form(""),
    mcp_keywords: str = Form(""),
    ai_deepseek_api_key: str = Form(""),
    ai_deepseek_base_url: str = Form(""),
    ai_deepseek_model: str = Form(""),
    ai_glm_api_key: str = Form(""),
    ai_glm_base_url: str = Form(""),
    ai_glm_model: str = Form(""),
    ai_openai_api_key: str = Form(""),
    ai_openai_base_url: str = Form(""),
    ai_openai_model: str = Form(""),
    ai_route_news: str = Form("deepseek"),
    ai_route_weather: str = Form("openai"),
    ai_route_schedule: str = Form("glm"),
    archive_database_url: str = Form(""),
    archive_enabled: str = Form("false"),
    archive_cron: str = Form("0 2 * * 1"),
    archive_run_time: str = Form(""),
    archive_run_weekday: str = Form(""),
    db: Session = Depends(get_db),
):
    redirect = require_admin(request)
    if redirect:
        return redirect
    crud.upsert_setting(db, "wechat_appid", wechat_appid)
    crud.upsert_setting(db, "wechat_secret", wechat_secret)
    crud.upsert_setting(db, "weather_geo_url", weather_geo_url)
    crud.upsert_setting(db, "weather_api_url", weather_api_url)
    crud.upsert_setting(db, "news_crawler_enabled", news_crawler_enabled)
    crud.upsert_setting(db, "news_crawler_cron", news_crawler_cron)
    crud.upsert_setting(db, "news_crawler_interval_minutes", news_crawler_interval_minutes)
    crud.upsert_setting(db, "news_crawler_limit", news_crawler_limit)
    crud.upsert_setting(db, "news_target_count", news_target_count)
    crud.upsert_setting(db, "news_dedupe_max_rounds", news_dedupe_max_rounds)
    crud.upsert_setting(db, "news_dedupe_max_candidates", news_dedupe_max_candidates)
    if news_run_time:
        crud.upsert_setting(db, "news_crawler_cron", build_weekly_cron(news_run_time, news_run_weekday or "daily"))
    else:
        crud.upsert_setting(db, "news_crawler_cron", news_crawler_cron)
    crud.upsert_setting(db, "news_source_rss", news_source_rss)
    crud.upsert_setting(db, "news_source_mcp", news_source_mcp)
    crud.upsert_setting(db, "news_source_feeds", news_source_feeds)
    crud.upsert_setting(db, "news_feed_urls", news_feed_urls)
    crud.upsert_setting(db, "news_source_mode", news_source_mode)
    crud.upsert_setting(db, "mcp_enabled", mcp_enabled)
    crud.upsert_setting(db, "mcp_base_url", mcp_base_url)
    crud.upsert_setting(db, "mcp_api_key", mcp_api_key)
    crud.upsert_setting(db, "mcp_keywords", mcp_keywords)
    crud.upsert_setting(db, "ai_deepseek_api_key", ai_deepseek_api_key)
    crud.upsert_setting(db, "ai_deepseek_base_url", ai_deepseek_base_url)
    crud.upsert_setting(db, "ai_deepseek_model", ai_deepseek_model)
    crud.upsert_setting(db, "ai_glm_api_key", ai_glm_api_key)
    crud.upsert_setting(db, "ai_glm_base_url", ai_glm_base_url)
    crud.upsert_setting(db, "ai_glm_model", ai_glm_model)
    crud.upsert_setting(db, "ai_openai_api_key", ai_openai_api_key)
    crud.upsert_setting(db, "ai_openai_base_url", ai_openai_base_url)
    crud.upsert_setting(db, "ai_openai_model", ai_openai_model)
    crud.upsert_setting(db, "ai_route_news", ai_route_news)
    crud.upsert_setting(db, "ai_route_weather", ai_route_weather)
    crud.upsert_setting(db, "ai_route_schedule", ai_route_schedule)
    if archive_database_url:
        crud.upsert_setting(db, "archive_database_url", archive_database_url)
    crud.upsert_setting(db, "archive_enabled", archive_enabled)
    if archive_run_time:
        crud.upsert_setting(db, "archive_cron", build_weekly_cron(archive_run_time, archive_run_weekday or "1"))
    else:
        crud.upsert_setting(db, "archive_cron", archive_cron)
    configure_news_scheduler(db)
    configure_archive_scheduler(db)
    return RedirectResponse(url="/admin/settings", status_code=status.HTTP_302_FOUND)


@app.get("/admin/users")
def admin_users(request: Request, q: str | None = None, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    users = crud.list_users_filtered(db, q)
    return templates.TemplateResponse(
        "admin_users.html",
        {"request": request, "users": users, "query": q or ""},
    )


@app.get("/admin/users/{user_id}")
def admin_user_detail(user_id: int, request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    user = crud.get_user_by_id(db, user_id)
    if not user:
        return RedirectResponse(url="/admin/users", status_code=status.HTTP_302_FOUND)
    return templates.TemplateResponse(
        "admin_user_detail.html",
        {
            "request": request,
            "user": user,
            "subscriptions": user.subscriptions,
            "message": "",
        },
    )


@app.post("/admin/users/{user_id}")
def admin_user_detail_post(
    user_id: int,
    request: Request,
    username: str = Form(...),
    location: str = Form(""),
    subscriptions: str = Form(""),
    is_active: str = Form("true"),
    reset_password: str = Form(""),
    action: str = Form("update"),
    db: Session = Depends(get_db),
):
    redirect = require_admin(request)
    if redirect:
        return redirect
    user = crud.get_user_by_id(db, user_id)
    if not user:
        return RedirectResponse(url="/admin/users", status_code=status.HTTP_302_FOUND)
    if action == "deactivate":
        crud.set_user_active(db, user, False)
        message = "用户已停用"
    else:
        updated_user = crud.update_user_profile(
            db,
            user,
            username=username,
            location=location,
            subscriptions=[tag.strip() for tag in subscriptions.split(",") if tag.strip()],
            email=user.email,
            phone=user.phone,
            avatar_url=user.avatar_url,
        )
        if reset_password:
            crud.reset_user_password(db, updated_user, reset_password)
        crud.set_user_active(db, updated_user, is_active.lower() == "true")
        message = "用户信息已更新"
    return templates.TemplateResponse(
        "admin_user_detail.html",
        {
            "request": request,
            "user": crud.get_user_by_id(db, user_id),
            "subscriptions": subscriptions,
            "message": message,
        },
    )


@app.get("/admin/tasks")
def admin_tasks(request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    runs = crud.list_task_runs(db, limit=20)
    return templates.TemplateResponse(
        "admin_tasks.html",
        {"request": request, "runs": runs, "message": ""},
    )


@app.get("/admin/scheduler")
def admin_scheduler(request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    scheduler_info = configure_news_scheduler(db)
    settings_map = {
        "news_crawler_enabled": _setting_value(db, "news_crawler_enabled", "false"),
        "news_crawler_cron": _setting_value(db, "news_crawler_cron", ""),
        "news_crawler_interval_minutes": _setting_value(db, "news_crawler_interval_minutes", "30"),
        "news_crawler_limit": _setting_value(db, "news_crawler_limit", "20"),
        "news_source_rss": _setting_value(db, "news_source_rss", "true"),
        "news_source_mcp": _setting_value(db, "news_source_mcp", "false"),
        "news_source_feeds": _setting_value(db, "news_source_feeds", "true"),
        "next_run_at": scheduler_info.get("next_run_at", ""),
        "scheduler_error": scheduler_info.get("error", ""),
    }
    news_time, news_weekday = parse_weekly_cron(settings_map["news_crawler_cron"])
    settings_map["news_run_time"] = news_time
    settings_map["news_run_weekday"] = news_weekday
    return templates.TemplateResponse(
        "admin_scheduler.html",
        {"request": request, "settings": settings_map},
    )


@app.post("/admin/scheduler")
def admin_scheduler_post(
    request: Request,
    news_crawler_enabled: str = Form("false"),
    news_crawler_cron: str = Form(""),
    news_crawler_interval_minutes: str = Form("30"),
    news_crawler_limit: str = Form("20"),
    news_source_rss: str = Form("true"),
    news_source_mcp: str = Form("false"),
    news_source_feeds: str = Form("true"),
    news_run_time: str = Form(""),
    news_run_weekday: str = Form(""),
    db: Session = Depends(get_db),
):
    redirect = require_admin(request)
    if redirect:
        return redirect
    crud.upsert_setting(db, "news_crawler_enabled", news_crawler_enabled)
    crud.upsert_setting(db, "news_crawler_cron", news_crawler_cron)
    crud.upsert_setting(db, "news_crawler_interval_minutes", news_crawler_interval_minutes)
    crud.upsert_setting(db, "news_crawler_limit", news_crawler_limit)
    if news_run_time:
        crud.upsert_setting(db, "news_crawler_cron", build_weekly_cron(news_run_time, news_run_weekday or "daily"))
    else:
        crud.upsert_setting(db, "news_crawler_cron", news_crawler_cron)
    crud.upsert_setting(db, "news_source_rss", news_source_rss)
    crud.upsert_setting(db, "news_source_mcp", news_source_mcp)
    crud.upsert_setting(db, "news_source_feeds", news_source_feeds)
    configure_news_scheduler(db)
    return RedirectResponse(url="/admin/scheduler", status_code=status.HTTP_302_FOUND)


@app.get("/admin/mcp")
def admin_mcp(request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    remotes = crud.list_mcp_remotes(db)
    locals_ = crud.list_mcp_locals(db)
    return templates.TemplateResponse(
        "admin_mcp.html",
        {"request": request, "remotes": remotes, "locals": locals_},
    )


@app.get("/admin/weather")
def admin_weather(request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    providers = crud.list_weather_providers(db)
    return templates.TemplateResponse(
        "admin_weather.html",
        {"request": request, "providers": providers},
    )


@app.get("/admin/news")
def admin_news(request: Request, db: Session = Depends(get_db)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    return templates.TemplateResponse(
        "admin_news.html",
        {"request": request, "message": ""},
    )


@app.post("/admin/news")
def admin_news_post(
    request: Request,
    title: str = Form(...),
    url: str = Form(...),
    source: str = Form(...),
    published_at: str = Form(...),
    tags: str = Form(""),
    content: str = Form(""),
    summary: str = Form(""),
    db: Session = Depends(get_db),
):
    redirect = require_admin(request)
    if redirect:
        return redirect
    final_summary = summary.strip() or news_summary_from_ai(db, title, content)
    crud.create_news_item(
        db,
        title=title,
        summary=final_summary,
        source=source,
        url=url,
        published_at=published_at,
        tags=[tag.strip() for tag in tags.split(",") if tag.strip()],
    )
    return templates.TemplateResponse(
        "admin_news.html",
        {"request": request, "message": "资讯已保存"},
    )


@app.post("/admin/api/tasks/news_crawl")
async def admin_task_news_crawl(
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    start = time.time()
    limit = int(payload.get("limit") or _setting_value(db, "news_crawler_limit", "20"))
    result = {}
    error_message = ""
    try:
        result = run_news_crawler(limit=limit, force_run=True)
        status_text = "success"
    except Exception as exc:
        status_text = "failed"
        error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    run = crud.create_task_run(
        db,
        task_type="news_crawl",
        status=status_text,
        duration_ms=duration_ms,
        error_message=error_message,
        log_excerpt=json.dumps(result, ensure_ascii=False),
        payload_json=json.dumps(payload, ensure_ascii=False),
    )
    return {"status": status_text, "duration_ms": duration_ms, "error_message": error_message, "log_excerpt": run.log_excerpt, "task_run_id": run.id}


@app.post("/admin/api/tasks/weather_refresh")
async def admin_task_weather_refresh(
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    start = time.time()
    location = payload.get("location", "")
    error_message = ""
    try:
        fetch_weather(db, location)
        status_text = "success"
    except Exception as exc:
        status_text = "failed"
        error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    run = crud.create_task_run(
        db,
        task_type="weather_refresh",
        status=status_text,
        duration_ms=duration_ms,
        error_message=error_message,
        log_excerpt="",
        payload_json=json.dumps(payload, ensure_ascii=False),
    )
    return {"status": status_text, "duration_ms": duration_ms, "error_message": error_message, "log_excerpt": run.log_excerpt, "task_run_id": run.id}


@app.post("/admin/api/tasks/schedule_analyze")
async def admin_task_schedule_analyze(
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    start = time.time()
    error_message = ""
    try:
        user_id = int(payload.get("user_id") or 0)
        plan = crud.get_latest_schedule(db, user_id) if user_id else None
        status_text = "success" if plan else "failed"
        if not plan:
            error_message = "未找到可分析的日程计划"
    except Exception as exc:
        status_text = "failed"
        error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    run = crud.create_task_run(
        db,
        task_type="schedule_analyze",
        status=status_text,
        duration_ms=duration_ms,
        error_message=error_message,
        log_excerpt="",
        payload_json=json.dumps(payload, ensure_ascii=False),
    )
    return {"status": status_text, "duration_ms": duration_ms, "error_message": error_message, "log_excerpt": run.log_excerpt, "task_run_id": run.id}


@app.post("/admin/api/archive/run")
async def admin_archive_run(request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    start = time.time()
    error_message = ""
    status_text = "success"
    result = {}
    try:
        result = run_archive_job(force_run=True)
    except Exception as exc:
        status_text = "failed"
        error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    run = crud.create_task_run(
        db,
        task_type="archive_manual",
        status=status_text,
        duration_ms=duration_ms,
        error_message=error_message,
        log_excerpt=json.dumps(result, ensure_ascii=False),
        payload_json=json.dumps({}, ensure_ascii=False),
    )
    return {"status": status_text, "duration_ms": duration_ms, "error_message": error_message, "log_excerpt": run.log_excerpt, "task_run_id": run.id}


@app.post("/admin/api/archive/test")
async def admin_archive_test(request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    archive_session = get_archive_session(db)
    if not archive_session:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="ARCHIVE_DB_NOT_CONFIGURED")
    try:
        archive_session.execute(text("SELECT 1"))
    finally:
        archive_session.close()
    return {"status": "success"}


@app.get("/admin/api/scheduler")
def admin_api_scheduler(request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    scheduler_info = configure_news_scheduler(db)
    return {
        "news_crawler_enabled": _setting_bool(db, "news_crawler_enabled", "false"),
        "cron": _setting_value(db, "news_crawler_cron", ""),
        "interval_minutes": int(_setting_value(db, "news_crawler_interval_minutes", "30") or 30),
        "limit": int(_setting_value(db, "news_crawler_limit", "20") or 20),
        "sources": {
            "rss": _setting_bool(db, "news_source_rss", "true"),
            "mcp": _setting_bool(db, "news_source_mcp", "false"),
            "feeds": _setting_bool(db, "news_source_feeds", "true"),
        },
        "next_run_at": scheduler_info.get("next_run_at", ""),
        "last_runs": [_task_run_to_dict(run) for run in crud.list_task_runs(db, limit=5)],
    }


@app.put("/admin/api/scheduler")
def admin_api_scheduler_put(payload: schemas.SchedulerConfig, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    crud.upsert_setting(db, "news_crawler_enabled", "true" if payload.news_crawler_enabled else "false")
    crud.upsert_setting(db, "news_crawler_cron", payload.cron)
    crud.upsert_setting(db, "news_crawler_interval_minutes", str(payload.interval_minutes))
    crud.upsert_setting(db, "news_crawler_limit", str(payload.limit))
    crud.upsert_setting(db, "news_source_rss", "true" if payload.sources.get("rss") else "false")
    crud.upsert_setting(db, "news_source_mcp", "true" if payload.sources.get("mcp") else "false")
    crud.upsert_setting(db, "news_source_feeds", "true" if payload.sources.get("feeds") else "false")
    scheduler_info = configure_news_scheduler(db)
    return {"status": "ok", "next_run_at": scheduler_info.get("next_run_at", "")}


@app.get("/admin/api/mcp/remotes")
def admin_api_mcp_remotes(request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    return [
        {
            "id": remote.id,
            "name": remote.name,
            "base_url": remote.base_url,
            "protocol": remote.protocol,
            "auth_type": remote.auth_type,
            "auth_value": remote.auth_value,
            "extra_config": json.loads(remote.extra_config or "{}"),
            "timeout_sec": remote.timeout_sec,
            "enabled": remote.enabled,
            "priority": remote.priority,
        }
        for remote in crud.list_mcp_remotes(db)
    ]


@app.post("/admin/api/mcp/remotes")
def admin_api_mcp_remote_create(
    payload: schemas.MCPRemoteConfigPayload, request: Request, db: Session = Depends(get_db)
):
    require_admin_json(request)
    remote = crud.create_mcp_remote(
        db,
        name=payload.name,
        base_url=payload.base_url,
        protocol=payload.protocol,
        auth_type=payload.auth_type,
        auth_value=payload.auth_value,
        extra_config=json.dumps(payload.extra_config, ensure_ascii=False),
        timeout_sec=payload.timeout_sec,
        enabled=payload.enabled,
        priority=payload.priority,
    )
    return {"id": remote.id}


@app.put("/admin/api/mcp/remotes/{remote_id}")
def admin_api_mcp_remote_update(
    remote_id: int,
    payload: schemas.MCPRemoteConfigPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    remote = crud.get_mcp_remote(db, remote_id)
    if not remote:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Remote not found")
    crud.update_mcp_remote(
        db,
        remote,
        name=payload.name,
        base_url=payload.base_url,
        protocol=payload.protocol,
        auth_type=payload.auth_type,
        auth_value=payload.auth_value,
        extra_config=json.dumps(payload.extra_config, ensure_ascii=False),
        timeout_sec=payload.timeout_sec,
        enabled=payload.enabled,
        priority=payload.priority,
    )
    return {"status": "ok"}


@app.delete("/admin/api/mcp/remotes/{remote_id}")
def admin_api_mcp_remote_delete(remote_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    remote = crud.get_mcp_remote(db, remote_id)
    if not remote:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Remote not found")
    crud.delete_mcp_remote(db, remote)
    return {"status": "ok"}


@app.post("/admin/api/mcp/remotes/{remote_id}/test")
def admin_api_mcp_remote_test(remote_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    remote = crud.get_mcp_remote(db, remote_id)
    if not remote:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Remote not found")
    start = time.time()
    error_message = ""
    status_text = "success"
    try:
        headers = {}
        if remote.auth_type in {"api_key", "token"} and remote.auth_value:
            headers["Authorization"] = f"Bearer {remote.auth_value}"
        response = httpx.post(remote.base_url, json={"query": "测试", "limit": 1}, headers=headers, timeout=remote.timeout_sec)
        response.raise_for_status()
    except Exception as exc:
        status_text = "failed"
        error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    return {"status": status_text, "duration_ms": duration_ms, "error_message": error_message}


@app.get("/admin/api/mcp/locals")
def admin_api_mcp_locals(request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    return [
        {
            "id": local.id,
            "name": local.name,
            "module_path": local.module_path,
            "command": local.command,
            "args": json.loads(local.args_json or "[]"),
            "env": json.loads(local.env_json or "{}"),
            "capabilities": json.loads(local.capabilities or "[]"),
            "schema": json.loads(local.schema or "{}"),
            "timeout_sec": local.timeout_sec,
            "enabled": local.enabled,
            "priority": local.priority,
        }
        for local in crud.list_mcp_locals(db)
    ]


@app.post("/admin/api/mcp/locals")
def admin_api_mcp_local_create(
    payload: schemas.MCPLocalPluginPayload, request: Request, db: Session = Depends(get_db)
):
    require_admin_json(request)
    local = crud.create_mcp_local(
        db,
        name=payload.name,
        module_path=payload.module_path,
        command=payload.command,
        args_json=json.dumps(payload.args, ensure_ascii=False),
        env_json=json.dumps(payload.env, ensure_ascii=False),
        capabilities=json.dumps(payload.capabilities, ensure_ascii=False),
        schema=json.dumps(payload.schema, ensure_ascii=False),
        timeout_sec=payload.timeout_sec,
        enabled=payload.enabled,
        priority=payload.priority,
    )
    return {"id": local.id}


@app.put("/admin/api/mcp/locals/{local_id}")
def admin_api_mcp_local_update(
    local_id: int,
    payload: schemas.MCPLocalPluginPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    local = crud.get_mcp_local(db, local_id)
    if not local:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Local plugin not found")
    crud.update_mcp_local(
        db,
        local,
        name=payload.name,
        module_path=payload.module_path,
        command=payload.command,
        args_json=json.dumps(payload.args, ensure_ascii=False),
        env_json=json.dumps(payload.env, ensure_ascii=False),
        capabilities=json.dumps(payload.capabilities, ensure_ascii=False),
        schema=json.dumps(payload.schema, ensure_ascii=False),
        timeout_sec=payload.timeout_sec,
        enabled=payload.enabled,
        priority=payload.priority,
    )
    return {"status": "ok"}


@app.delete("/admin/api/mcp/locals/{local_id}")
def admin_api_mcp_local_delete(local_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    local = crud.get_mcp_local(db, local_id)
    if not local:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Local plugin not found")
    crud.delete_mcp_local(db, local)
    return {"status": "ok"}


@app.post("/admin/api/mcp/locals/{local_id}/test")
def admin_api_mcp_local_test(local_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    local = crud.get_mcp_local(db, local_id)
    if not local:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Local plugin not found")
    start = time.time()
    error_message = ""
    status_text = "success"
    try:
        if local.command:
            args = [local.command] + json.loads(local.args_json or "[]")
            env = os.environ.copy()
            env.update(json.loads(local.env_json or "{}"))
            timeout_sec = local.timeout_sec or 10
            process = subprocess.Popen(
                args,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                env=env,
                text=True,
            )
            try:
                stdout, stderr = process.communicate(timeout=timeout_sec)
            except subprocess.TimeoutExpired:
                process.terminate()
                try:
                    process.wait(timeout=1)
                except subprocess.TimeoutExpired:
                    process.kill()
                    process.wait(timeout=1)
                error_message = "命令在超时内未退出，可能为常驻服务"
            else:
                if process.returncode != 0:
                    status_text = "failed"
                    output = (stderr or stdout or "").strip()
                    error_message = output[:200] if output else f"命令退出码 {process.returncode}"
                else:
                    stderr = (stderr or "").strip()
                    if stderr:
                        error_message = stderr[:200]
        else:
            test_plugin(local.module_path)
    except MCPPluginError as exc:
        status_text = "failed"
        error_message = str(exc)
    except subprocess.TimeoutExpired:
        status_text = "failed"
        error_message = "命令执行超时"
    except (json.JSONDecodeError, subprocess.SubprocessError, FileNotFoundError) as exc:
        status_text = "failed"
        error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    return {"status": status_text, "duration_ms": duration_ms, "error_message": error_message}


@app.get("/admin/api/weather/providers")
def admin_api_weather_providers(request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    return [
        {
            "id": provider.id,
            "name": provider.name,
            "provider_type": provider.provider_type,
            "base_url": provider.base_url,
            "api_key": provider.api_key,
            "timeout_sec": provider.timeout_sec,
            "enabled": provider.enabled,
            "priority": provider.priority,
            "extra_config": json.loads(provider.extra_config or "{}"),
        }
        for provider in crud.list_weather_providers(db)
    ]


@app.post("/admin/api/weather/providers")
def admin_api_weather_provider_create(
    payload: schemas.WeatherProviderPayload, request: Request, db: Session = Depends(get_db)
):
    require_admin_json(request)
    provider = crud.create_weather_provider(
        db,
        name=payload.name,
        provider_type=payload.provider_type,
        base_url=payload.base_url,
        api_key=payload.api_key,
        timeout_sec=payload.timeout_sec,
        enabled=payload.enabled,
        priority=payload.priority,
        extra_config=json.dumps(payload.extra_config, ensure_ascii=False),
    )
    return {"id": provider.id}


@app.put("/admin/api/weather/providers/{provider_id}")
def admin_api_weather_provider_update(
    provider_id: int,
    payload: schemas.WeatherProviderPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    provider = crud.get_weather_provider(db, provider_id)
    if not provider:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found")
    crud.update_weather_provider(
        db,
        provider,
        name=payload.name,
        provider_type=payload.provider_type,
        base_url=payload.base_url,
        api_key=payload.api_key,
        timeout_sec=payload.timeout_sec,
        enabled=payload.enabled,
        priority=payload.priority,
        extra_config=json.dumps(payload.extra_config, ensure_ascii=False),
    )
    return {"status": "ok"}


@app.delete("/admin/api/weather/providers/{provider_id}")
def admin_api_weather_provider_delete(provider_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    provider = crud.get_weather_provider(db, provider_id)
    if not provider:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found")
    crud.delete_weather_provider(db, provider)
    return {"status": "ok"}


@app.post("/admin/api/weather/providers/{provider_id}/test")
def admin_api_weather_provider_test(provider_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    provider = crud.get_weather_provider(db, provider_id)
    if not provider:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found")
    start = time.time()
    status_text = "success"
    error_message = ""
    test_location = "上海"
    try:
        extra = json.loads(provider.extra_config or "{}")
        test_location = extra.get("test_location", test_location)
    except json.JSONDecodeError:
        error_message = "extra_config JSON 格式错误"
        status_text = "failed"
    if status_text == "success":
        try:
            extra = json.loads(provider.extra_config or "{}")
            if provider.provider_type == "open-meteo":
                _fetch_open_meteo(
                    test_location,
                    geo_url=extra.get("geo_url", settings.weather_geo_url),
                    api_url=provider.base_url,
                    timeout=float(provider.timeout_sec or 5),
                )
            elif provider.provider_type == "gaode":
                _fetch_gaode(
                    test_location,
                    base_url=provider.base_url,
                    api_key=provider.api_key,
                    timeout=float(provider.timeout_sec or 5),
                )
            else:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported provider")
        except Exception as exc:
            status_text = "failed"
            error_message = str(exc)
    duration_ms = int((time.time() - start) * 1000)
    return {"status": status_text, "provider": provider.provider_type, "duration_ms": duration_ms, "error_message": error_message, "failover_count": 0}


@app.get("/admin/api/users")
def admin_api_users(request: Request, q: str | None = None, db: Session = Depends(get_db)):
    require_admin_json(request)
    users = crud.list_users_filtered(db, q)
    return [
        {
            "id": user.id,
            "username": user.username,
            "location": user.location,
            "subscriptions": [tag for tag in user.subscriptions.split(",") if tag],
            "is_active": user.is_active,
        }
        for user in users
    ]


@app.get("/admin/api/users/{user_id}")
def admin_api_user_detail(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    user = crud.get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return {
        "id": user.id,
        "username": user.username,
        "location": user.location,
        "subscriptions": [tag for tag in user.subscriptions.split(",") if tag],
        "is_active": user.is_active,
    }


@app.put("/admin/api/users/{user_id}")
def admin_api_user_update(
    user_id: int,
    payload: schemas.UserAdminPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    user = crud.get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    updated_user = crud.update_user_profile(
        db,
        user,
        username=payload.username,
        location=payload.location,
        subscriptions=payload.subscriptions,
        email=user.email,
        phone=user.phone,
        avatar_url=user.avatar_url,
    )
    crud.set_user_active(db, updated_user, payload.is_active)
    return {"status": "ok"}


@app.delete("/admin/api/users/{user_id}")
def admin_api_user_delete(user_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin_json(request)
    user = crud.get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    crud.set_user_active(db, user, False)
    return {"status": "ok"}


@app.post("/admin/api/users/{user_id}/reset_password")
def admin_api_user_reset_password(
    user_id: int,
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    require_admin_json(request)
    user = crud.get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    new_password = payload.get("new_password", "")
    if not new_password:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="new_password required")
    crud.reset_user_password(db, user, new_password)
    return {"status": "ok"}
